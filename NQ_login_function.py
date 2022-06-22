import re,json, openpyxl
import time, random
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException, WebDriverException
from selenium.webdriver.support import expected_conditions as EC
from random import choice
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell


import pathlib
from pathlib import Path
import os
from sys import platform
import NQ_function


chrome_options = webdriver.ChromeOptions()

class objects:
    now = datetime.now()
    year = now.strftime("%Y")
    month = now.strftime("%m")
    day = now.strftime("%d")
    time1 = now.strftime("%H:%M:%S")
    date_time = now.strftime("%Y/%m/%d, %H:%M:%S")
    date_id = date_time.replace("/", "").replace(", ", "").replace(":", "")[2:]
    testcase_pass = "Test case status: pass"
    testcase_fail = "Test case status: fail"

if platform == "linux" or platform == "linux2":
    local_path = "/home/oem/groupware-auto-test"
    json_file = local_path + "/NQ_selenium.json"
    with open(json_file) as json_data_file:
        data = json.load(json_data_file)
    driver = webdriver.Chrome("/usr/bin/chromedriver")
    log_folder = "/Log/"
    log_testcase = "/Log/"
    file_upload = local_path+"/Attachment/quynh1@meo.qa.hanbiro.net_2323_1608785342.837046.eml"
    file_zip_upload = local_path +"/Attachment/sent_mail.zip"
    execution_log = local_path + log_folder + "execution_log_" + str(objects.date_id) + ".txt"
    # fail_log = execution_log.replace("execution_log_", "fail_log_")
    # error_log = execution_log.replace("execution_log_", "error_log_")
    testcase_log = local_path + log_testcase + "NQuynh_TestcaseAllmenu_" + str(objects.date_id) + ".xlsx"
else :
    local_path = os.path.dirname(Path(__file__).absolute())
    json_file = local_path + "\\NQ_selenium.json"
    with open(json_file) as json_data_file:
        data = json.load(json_data_file)
    driver = webdriver.Chrome(local_path + "\\chromedriver.exe")
    log_folder = "\\Log\\"
    log_testcase = "\\Log\\"
    file_upload = local_path +"/Attachment/quynh1@meo.qa.hanbiro.net_2323_1608785342.837046.eml"
    file_zip_upload = local_path +"/Attachment/sent_mail.zip"
    execution_log = local_path + log_folder + "execution_log_" + str(objects.date_id) + ".txt"
    # fail_log = execution_log.replace("execution_log_", "fail_log_")
    # error_log = execution_log.replace("execution_log_", "error_log_")
    testcase_log = local_path + log_testcase + "NQuynh_TestcaseAllmenu_" + str(objects.date_id) + ".xlsx"

'''# create log file of fail test case
open(execution_log, "x").close()

# create log file of fail test case
open(fail_log, "x").close()

# create log file of fail test case
open(error_log, "x").close()'''

# excel file

logs = [testcase_log]
for log in logs:
    if ".txt" in log:
        open(log, "x").close()
    else:
        wb = Workbook()
        myFill = PatternFill(start_color='adc5e7',
                   end_color='adc5e7',
                   fill_type='solid',)
        font = Font(name='Calibri',
                    size=11 ,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
        ws = wb.active

        ws.cell(row=1, column=1).value= "Menu"
        ws.cell(row=1, column=2).value = "Sub-Menu"
        ws.cell(row=1, column=3).value = "Test Case Name"
        ws.cell(row=1, column=4).value = "Status"
        ws.cell(row=1, column=5).value = "Description"
        ws.cell(row=1, column=6).value = "Date"
        ws.cell(row=1, column=7).value = "Tester"
        # color 
        ws.cell(row=1, column=1).fill = myFill
        ws.cell(row=1, column=2).fill = myFill
        ws.cell(row=1, column=3).fill = myFill
        ws.cell(row=1, column=4).fill = myFill
        ws.cell(row=1, column=5).fill = myFill
        ws.cell(row=1, column=6).fill = myFill
        ws.cell(row=1, column=7).fill = myFill
        # font
        ws.cell(row=1, column=1).font = Font(bold=True)
        ws.cell(row=1, column=2).font = Font(bold=True)
        ws.cell(row=1, column=3).font = Font(bold=True)
        ws.cell(row=1, column=4).font = Font(bold=True)
        ws.cell(row=1, column=5).font = Font(bold=True)
        ws.cell(row=1, column=6).font = Font(bold=True)
        ws.cell(row=1, column=7).font = Font(bold=True)

        wb.save(log)

def Logging(*messages):
    msg = str(" ".join(list(messages))) 
    print(msg)
    log_msg = open(execution_log, "a")
    written_msg = str(msg).encode(encoding="ascii",errors="ignore")
    log_msg.write(str(written_msg) + "\n")
    log_msg.close()

def TesCase_LogResult(menu, sub_menu, testcase, status, description, tester):
    Logging(description)

    # if status == "Pass":
    #     print(objects.testcase_pass)
    # else:
    #     print(objects.testcase_fail)

    wb = openpyxl.load_workbook(testcase_log)
    current_sheet = wb.active
    start_row = len(list(current_sheet.rows)) + 1

    current_sheet.cell(row=start_row, column=1).value = menu
    current_sheet.cell(row=start_row, column=2).value = sub_menu
    current_sheet.cell(row=start_row, column=3).value = testcase
    current_sheet.cell(row=start_row, column=4).value = status
    current_sheet.cell(row=start_row, column=5).value = description
    current_sheet.cell(row=start_row, column=6).value = objects.date_time
    current_sheet.cell(row=start_row, column=7).value = tester

    # Apply color for status: Pass/Fail
    passFill = PatternFill(start_color='b6d7a8',
                   end_color='b6d7a8',
                   fill_type='solid',)
    failFill = PatternFill(start_color='ea9999',
                   end_color='ea9999',
                   fill_type='solid')
    if status == "Pass":
        print(objects.testcase_pass)
        current_sheet.cell(row=start_row, column=4).fill = passFill
    else:
        print(objects.testcase_fail)
        current_sheet.cell(row=start_row, column=4).fill = failFill
    wb.save(testcase_log)

def ValidateFailResultAndSystem(fail_msg):
    print(fail_msg)
    append_fail_result = open(fail_log, "a")
    append_fail_result.write("[FAILED TEST CASE] " + str(fail_msg) + "\n")
    append_fail_result.close()



# ''' Test Link '''
# TESTLINK_API_PYTHON_SERVER_URL = 'http://qa1.hanbiro.net/testlink/lib/api/xmlrpc/v1/xmlrpc.php'
# TESTLINK_API_PYTHON_DEVKEY = 'b52edc5fc95e4fc0149d47da85e05f18'

# tls = testlink.TestLinkHelper(TESTLINK_API_PYTHON_SERVER_URL, TESTLINK_API_PYTHON_DEVKEY).connect(testlink.TestlinkAPIClient)

# def TestlinkResult_Pass(external_id):
#     tls.reportTCResult(testcaseexternalid=external_id, testplanid=7378, buildname="V3.8.33", status='p', notes='Test Case [' + external_id + '] passed')

# def TestlinkResult_Fail(external_id):
#     tls.reportTCResult(testcaseexternalid=external_id, testplanid=7378, buildname="V3.8.33", status='f', notes='Test Case [' + external_id + '] failed')

# start login page
driver.implicitly_wait(5)
driver.set_window_size(1024, 600)
driver.maximize_window()

def access_qa(domain_name):
    driver.get(domain_name)
    Logging("- Access login page")
    time.sleep(5)

    username = driver.find_element_by_id(data["user_name"])
    username.send_keys(data["hanbiro_user_1"])
    Logging("- Input user ID")
    frame_element = driver.find_element_by_id(data["frameelement"])
    driver.switch_to.frame(frame_element)
    password = driver.find_element_by_id("p")
    password.send_keys(data["hanbiro_password"])
    Logging("- Input user password")
    driver.switch_to.default_content()
    submit = driver.find_element_by_id("btn-log")
    time.sleep(1)
    submit.click()
    Logging("- Click button Login")
    time.sleep(5)

def access_global3(domain_name):
    driver.get(domain_name)
    Logging("- Access login page")
    time.sleep(5)

    userID = driver.find_element_by_name("gw_id")
    #userID.send_keys("luu")
    userID.send_keys(data["hanbiro_user_1"])
    print("- Input user ID")
    #add_data_in_excel(param_excel["checkin"],"p","Input reason late")
    password = driver.find_element_by_name("gw_pass")
    #password.send_keys("matkhau1!")
    password.send_keys(data["hanbiro_password"])
    print("- Input user password")
    password.send_keys(Keys.ENTER)
    print("- Click button Sign in")
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, data["new_comanage"]["notify"])))
    print("=> Log in successfully")


# close server popup
def close_server_popup():
    try:
        time.sleep(5)
        modal_bootbox = driver.find_element_by_xpath(data["close_popup"])
        modal_bootbox.click()
        Logging("- Close server pop up successfully")
    except WebDriverException:
        Logging("- Server pop up didn't show up")
    
    time.sleep(3)




